"""
Provenance Blockchain Metrics Scraper
Scrapes https://provenance.io/pulse daily at 11:55 PM ET.
Captures all 4 time periods: 24h, 1W, 1M, 3M.
Writes to data.json (GitHub Pages) and provenance_metrics.xlsx.

Fields per period: loan_amount_funded, loans_funded, loan_amount_paid, loans_paid
Plus: total_participants (from 1W view, it's a cumulative stat)
"""

import asyncio
import json
import os
from datetime import datetime
import pytz
from playwright.async_api import async_playwright
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

EXCEL_FILE = "provenance_metrics.xlsx"
JSON_FILE  = "data.json"
URL        = "https://provenance.io/pulse"

# Tab definitions: (data_key_prefix, button_text, funded_label, loans_funded_label, paid_label, loans_paid_label)
TABS = [
    {
        "prefix":         "1w",
        "btn":            None,          # default view, no click needed
        "funded_label":   "Week's Loan Amount Funded",
        "funded_ct":      "Week's Loans Funded",
        "paid_label":     "Week's Loan Amount Paid",
        "paid_ct":        "Week's Loans Paid",
    },
    {
        "prefix":         "24h",
        "btn":            "24h",
        "funded_label":   "Today's Loan Amount Funded",
        "funded_ct":      "Today's Loans Funded",
        "paid_label":     "Today's Loan Amount Paid",
        "paid_ct":        "Today's Loans Paid",
    },
    {
        "prefix":         "1m",
        "btn":            "1m",
        "funded_label":   "Month's Loan Amount Funded",
        "funded_ct":      "Month's Loans Funded",
        "paid_label":     "Month's Loan Amount Paid",
        "paid_ct":        "Month's Loans Paid",
    },
    {
        "prefix":         "3m",
        "btn":            "3m",
        "funded_label":   "3 Months Loan Amount Funded",
        "funded_ct":      "3 Months Loans Funded",
        "paid_label":     "3 Months Loan Amount Paid",
        "paid_ct":        "3 Months Loans Paid",
    },
]

HEADERS = [
    "Date",
    # 24h
    "Loan Amt Funded (24h)", "Loans Funded (24h)", "Loan Amt Paid (24h)", "Loans Paid (24h)",
    # 1W
    "Loan Amt Funded (1W)",  "Loans Funded (1W)",  "Loan Amt Paid (1W)",  "Loans Paid (1W)",
    # 1M
    "Loan Amt Funded (1M)",  "Loans Funded (1M)",  "Loan Amt Paid (1M)",  "Loans Paid (1M)",
    # 3M
    "Loan Amt Funded (3M)",  "Loans Funded (3M)",  "Loan Amt Paid (3M)",  "Loans Paid (3M)",
    # All
    "Total Participants",
]


def extract_from_text(text: str, label: str) -> str:
    """
    Page card format:
        Label
        i          ← tooltip icon
        $value     ← what we want
        delta
        (pct%)
        Period
    """
    lines = [l.strip() for l in text.split("\n") if l.strip()]
    for i, line in enumerate(lines):
        if line == label and i + 2 < len(lines):
            return lines[i + 2]
    return "N/A"


async def click_tab(page, btn_text: str):
    """Click the SECOND matching tab button (second section on the page)."""
    try:
        buttons = page.locator(f"button.pulse-pill:has-text('{btn_text}')")
        count = await buttons.count()
        print(f"  Found {count} '{btn_text}' button(s)")
        await buttons.nth(1 if count >= 2 else 0).click()
        await page.wait_for_timeout(3000)
        print(f"  Clicked '{btn_text}' successfully")
    except Exception as e:
        print(f"  Could not click '{btn_text}': {e}")


async def scrape_metrics():
    metrics = {}

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=True)
        context = await browser.new_context(
            user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                       "(KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
        )
        page = await context.new_page()

        print(f"Loading {URL}...")
        await page.goto(URL, wait_until="networkidle", timeout=60000)
        await page.wait_for_timeout(4000)

        for tab in TABS:
            prefix = tab["prefix"]

            # Click tab if needed (1W is default, no click)
            if tab["btn"]:
                print(f"\nSwitching to {prefix} tab...")
                await click_tab(page, tab["btn"])
            else:
                print(f"\nScraping {prefix} (default view)...")

            text = await page.inner_text("body")

            metrics[f"loan_amount_funded_{prefix}"] = extract_from_text(text, tab["funded_label"])
            metrics[f"loans_funded_{prefix}"]        = extract_from_text(text, tab["funded_ct"])
            metrics[f"loan_amount_paid_{prefix}"]    = extract_from_text(text, tab["paid_label"])
            metrics[f"loans_paid_{prefix}"]          = extract_from_text(text, tab["paid_ct"])

            print(f"  Loan Amount Funded : {metrics[f'loan_amount_funded_{prefix}']}")
            print(f"  Loans Funded       : {metrics[f'loans_funded_{prefix}']}")
            print(f"  Loan Amount Paid   : {metrics[f'loan_amount_paid_{prefix}']}")
            print(f"  Loans Paid         : {metrics[f'loans_paid_{prefix}']}")

            # Grab total_participants from 1W view (cumulative, stable)
            if prefix == "1w":
                metrics["total_participants"] = extract_from_text(text, "Total Participants")
                print(f"  Total Participants : {metrics['total_participants']}")

        await browser.close()

    return metrics


# ── JSON ──────────────────────────────────────────────────────────────────────

def update_json(date_str: str, metrics: dict):
    data = []
    if os.path.exists(JSON_FILE):
        with open(JSON_FILE, "r") as f:
            try:
                data = json.load(f)
            except json.JSONDecodeError:
                data = []

    data.append({"date": date_str, **metrics})

    with open(JSON_FILE, "w") as f:
        json.dump(data, f, indent=2)
    print(f"\ndata.json updated ({len(data)} total records)")


# ── Excel ─────────────────────────────────────────────────────────────────────

def setup_workbook():
    wb = Workbook()
    ws = wb.active
    ws.title = "Provenance Metrics"

    header_font  = Font(name="Arial", bold=True, color="FFFFFF")
    header_fill  = PatternFill("solid", start_color="1F4E79")
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border  = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"),  bottom=Side(style="thin")
    )

    ws.append(HEADERS)
    for col_idx in range(1, len(HEADERS) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = center_align
        cell.border    = thin_border

    for i in range(1, len(HEADERS) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 24

    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"
    wb.save(EXCEL_FILE)
    print(f"Created new workbook: {EXCEL_FILE}")


def append_excel_row(date_str: str, metrics: dict):
    if not os.path.exists(EXCEL_FILE):
        setup_workbook()

    wb = load_workbook(EXCEL_FILE)
    ws = wb.active

    row_data = [
        date_str,
        metrics.get("loan_amount_funded_24h", "N/A"), metrics.get("loans_funded_24h",        "N/A"),
        metrics.get("loan_amount_paid_24h",   "N/A"), metrics.get("loans_paid_24h",           "N/A"),
        metrics.get("loan_amount_funded_1w",  "N/A"), metrics.get("loans_funded_1w",          "N/A"),
        metrics.get("loan_amount_paid_1w",    "N/A"), metrics.get("loans_paid_1w",            "N/A"),
        metrics.get("loan_amount_funded_1m",  "N/A"), metrics.get("loans_funded_1m",          "N/A"),
        metrics.get("loan_amount_paid_1m",    "N/A"), metrics.get("loans_paid_1m",            "N/A"),
        metrics.get("loan_amount_funded_3m",  "N/A"), metrics.get("loans_funded_3m",          "N/A"),
        metrics.get("loan_amount_paid_3m",    "N/A"), metrics.get("loans_paid_3m",            "N/A"),
        metrics.get("total_participants",     "N/A"),
    ]

    next_row     = ws.max_row + 1
    fill         = PatternFill("solid", start_color="D6E4F0" if next_row % 2 == 0 else "FFFFFF")
    thin_border  = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"),  bottom=Side(style="thin")
    )
    center_align = Alignment(horizontal="center", vertical="center")

    ws.append(row_data)
    for col_idx in range(1, len(row_data) + 1):
        cell           = ws.cell(row=next_row, column=col_idx)
        cell.font      = Font(name="Arial")
        cell.fill      = fill
        cell.border    = thin_border
        cell.alignment = center_align

    wb.save(EXCEL_FILE)
    print(f"Excel row {next_row} saved.")


# ── Main ──────────────────────────────────────────────────────────────────────

async def main():
    et       = pytz.timezone("America/New_York")
    now_et   = datetime.now(et)
    date_str = now_et.strftime("%Y-%m-%d %I:%M %p ET")

    print(f"Starting scrape at {date_str}")
    metrics = await scrape_metrics()
    update_json(date_str, metrics)
    append_excel_row(date_str, metrics)
    print("All done.")


if __name__ == "__main__":
    asyncio.run(main())
